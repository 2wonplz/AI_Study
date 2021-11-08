from openpyxl import load_workbook

wb = load_workbook("c.xlsx")
ws = wb.active

for x in range(1, 11):
    for y in range(1, 11):
        print(ws.cell(x, y).value, end=" ")
    print()

wb.close()