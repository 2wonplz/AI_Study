from openpyxl import Workbook, load_workbook

'''
Workbook : 엑셀 파일이 없을때 생성
load_workbook : 엑셀 파일이 있을때 불러옴
'''


class MExcel:
    def __init__(self):
        pass

    def save(self, a, b, c):
        wb = load_workbook("mexcel.xlsx")
        ws = wb.active
        ws.append([a, b, c])
        wb.save("mexcel.xlsx")
        wb.close()

    def load(self):
        wb = load_workbook("mexcel.xlsx")
        ws = wb.active

        for row in ws.iter_rows():
            for cell in row:
                print(cell.value, end=' ')
            print()

        wb.save("mexcel.xlsx")
        wb.close()

    def create(self):
        wb = Workbook()
        wb.save("mexcel.xlsx")
        wb.close()
