from openpyxl import *


class Grade:
    def __init__(self):
        pass

    def save(self, a, b, c):
        wb = load_workbook("a.xlsx")
        ws = wb.active
        ws.append([a, b, c, int(a)+int(b)+int(c), (int(a)+int(b)+int(c))/3])
        wb.save("a.xlsx")
        wb.close()

    def load(self):
        wb = load_workbook("a.xlsx")
        ws = wb.active

        for row in ws.iter_rows():
            for cell in row:
                print(cell.value, end=' ')
            print()

        wb.save("a.xlsx")
        wb.close()

    def create(self):
        wb = Workbook()
        wb.save("a.xlsx")
        wb.close()
