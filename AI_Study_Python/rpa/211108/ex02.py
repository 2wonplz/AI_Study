from openpyxl import Workbook

wb = Workbook()

ws1 = wb.create_sheet("newSheet1")
ws2 = wb.create_sheet("newSheet2", 2)

ws1.sheet_properties.tabColor = "ff66ff"
ws2.sheet_properties.tabColor = "6cd431"

ws1['A1'] = 'test'
ws2['B1'] = 'ws2222'

for i in range(1, 10):
    ws1['A' + str(i)] = 'test' + str(i)

target = wb.copy_worksheet(ws1)
target.title = 'copied sheet'

wb.save('b.xlsx')
wb.close()
