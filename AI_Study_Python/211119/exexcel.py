from openpyxl import *


def makebasic():
    wb = Workbook()
    ws = wb.active

    scores = [
        (1, 10, 8, 5, 14, 26, 12),
        (2, 7, 3, 7, 15, 24, 18),
        (3, 9, 5, 8, 8, 12, 4),
        (4, 7, 8, 7, 17, 21, 18),
        (5, 7, 8, 7, 16, 25, 15),
        (6, 3, 5, 8, 8, 17, 0),
        (7, 4, 9, 10, 16, 27, 18),
        (8, 6, 6, 6, 15, 19, 17),
        (9, 10, 10, 9, 19, 30, 19),
        (10, 9, 8, 8, 20, 25, 20)
    ]

    ws.append(["학번", "출석", "퀴즈1", "퀴즈2", "중간고사", "기말고사", "프로젝트"])
    for s in scores:
        ws.append(s)

    wb.save('score.xlsx')
    wb.close()


def maketen():
    wb = load_workbook('score.xlsx')
    ws = wb.active

    for idx, cell in enumerate(ws['D']):
        if idx == 0:
            continue
        else:
            cell.value = 10

    wb.save('score.xlsx')
    wb.close()


def maketotal():
    wb = load_workbook('score.xlsx')
    ws = wb.active

    ws['H1'] = '총점'
    for row in ws.iter_rows(min_row=2, min_col=2):
        sum = 0
        for cell in row[0:6]:
            sum += cell.value
        row[6].value = sum

    wb.save('score.xlsx')
    wb.close()


def makegrade():
    wb = load_workbook('score.xlsx')
    ws = wb.active

    ws['I1'] = '성적'

    try:
        for col in ws.iter_cols(min_col=8, min_row=2, max_col=9):
            for idx, cell in enumerate(col, start=2):
                if int(cell.value) >= 90:
                    ws['I' + str(idx)] = 'A'
                elif cell.value >= 80:
                    ws['I' + str(idx)] = 'B'
                elif cell.value >= 70:
                    ws['I' + str(idx)] = 'C'
                elif cell.value >= 60:
                    ws['I' + str(idx)] = 'D'
                else:
                    ws['I' + str(idx)] = 'F'
    except Exception as e:
        print(e)

    wb.save('score.xlsx')
    wb.close()
