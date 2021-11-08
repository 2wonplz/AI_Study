from openpyxl import Workbook
from random import randint

wb = Workbook()
ws = wb.active
ws.title = '제목'

ws['A1'] = 10
ws['A2'] = 20
ws['A3'] = 30

ws['B1'] = 40
ws['B2'] = 50
ws['B3'] = 60

print(ws['A2'])
print(ws['A2'].value)
print(ws['B3'].value)

ws.cell(column=1, row=1).value = 100
ws.cell(column=2, row=1).value = 200

ws.cell(2, 1, 300)

index = 1
for x in range(1, 11):
    for y in range(1, 11):
        ws.cell(x, y, randint(1, 100))
        index += 1

wb.save("c.xlsx")
wb.close()
