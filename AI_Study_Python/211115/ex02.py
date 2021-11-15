from openpyxl import load_workbook, Workbook
from openpyxl.chart import *

wb = Workbook()
ws = wb.active

for x in range(1, 11):
    for y in range(1, 11):
        ws.cell(x, y, x * y)

wb.save('sample.xlsx')
wb.close()


def insert_delete_files():
    wb = load_workbook('sample.xlsx')
    ws = wb.active

    ws.insert_rows(2)
    ws.insert_cols(2)
    ws.insert_cols(5, 3)
    # wb.save('sample_insert.xlsx')
    # wb.close()

    ws.delete_rows(5)
    ws.delete_cols(3)
    wb.save('sample_delete.xlsx')


def barmakeFile():
    wb = load_workbook('sample.xlsx')
    ws = wb.active

    bar_value = Reference(ws, 1, 1, 10, 10)
    bar_char = BarChart()
    bar_char.add_data(bar_value)
    ws.add_chart(bar_char, "G10")

    wb.save('sample_chart.xlsx')
    wb.close()


# insert_delete_files()
barmakeFile()
